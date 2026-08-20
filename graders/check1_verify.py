#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

REQUIRED = {"Raw Data", "Destination Summary", "Monthly Summary", "Key Metrics"}


def verify(output: Path, input_dir: Path) -> dict:
    checks, errors = {}, []
    report, workbook = output / "report.txt", output / "analysis.xlsx"
    checks["exact_deliverables"] = output.exists() and {p.name for p in output.iterdir()} == {"report.txt", "analysis.xlsx"}
    checks["report_present"] = report.is_file() and len(report.read_text(encoding="utf-8", errors="replace").strip()) >= 200
    checks["workbook_present"] = workbook.is_file()
    if openpyxl is None:
        return {"pass": False, "score": 0, "checks": checks, "errors": ["openpyxl is not installed"]}
    if not workbook.is_file():
        return {"pass": False, "score": 10 if checks["report_present"] else 0, "checks": checks, "errors": ["analysis.xlsx missing"]}
    try:
        wb = openpyxl.load_workbook(workbook, data_only=False)
    except Exception as exc:
        return {"pass": False, "score": 10, "checks": checks, "errors": [f"workbook open failed: {type(exc).__name__}"]}
    checks["required_sheets"] = REQUIRED.issubset(wb.sheetnames)
    with (input_dir / "destination_performance.csv").open(newline="", encoding="utf-8") as handle:
        source = list(csv.reader(handle))
    if "Raw Data" in wb.sheetnames:
        ws = wb["Raw Data"]
        raw = [[ws.cell(r, c).value for c in range(1, len(source[0]) + 1)] for r in range(1, len(source) + 1)]
        checks["raw_row_count"] = ws.max_row >= len(source)
        checks["raw_headers"] = [str(x) if x is not None else "" for x in raw[0]] == source[0]
        expected = [[str(v) for v in row[:3]] for row in source[1:]]
        actual = [[str(v) for v in row[:3]] for row in raw[1:]]
        checks["source_preserved"] = actual == expected
    else:
        checks.update({"raw_row_count": False, "raw_headers": False, "source_preserved": False})
    checks["destination_summary_nontrivial"] = "Destination Summary" in wb.sheetnames and wb["Destination Summary"].max_row >= 9 and wb["Destination Summary"].max_column >= 3
    checks["monthly_summary_nontrivial"] = "Monthly Summary" in wb.sheetnames and wb["Monthly Summary"].max_row >= 13 and wb["Monthly Summary"].max_column >= 3
    formulas = [cell.value for ws in wb.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    checks["reasonable_formula_count"] = len(formulas) >= 8
    checks["no_broken_formula_refs"] = not any("#REF!" in value.upper() for value in formulas)
    weights = {"exact_deliverables": 10, "report_present": 10, "workbook_present": 10,
               "required_sheets": 10, "raw_row_count": 10, "raw_headers": 10,
               "source_preserved": 15, "destination_summary_nontrivial": 5,
               "monthly_summary_nontrivial": 5, "reasonable_formula_count": 10,
               "no_broken_formula_refs": 5}
    score = sum(weight for name, weight in weights.items() if checks.get(name))
    passed = score >= 85 and all(checks.get(k) for k in ("exact_deliverables", "required_sheets", "source_preserved", "no_broken_formula_refs"))
    return {"pass": passed, "score": score, "checks": checks, "errors": errors, "formula_count": len(formulas)}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("output", type=Path)
    parser.add_argument("--input", type=Path, required=True)
    args = parser.parse_args()
    result = verify(args.output, args.input)
    print(json.dumps(result, indent=2, sort_keys=True))
    raise SystemExit(0 if result["pass"] else 1)


if __name__ == "__main__":
    main()

