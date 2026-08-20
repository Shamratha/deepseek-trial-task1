#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


def verify(output: Path, input_dir: Path) -> dict:
    workbook = output / "visual_brief.xlsx"
    checks = {"exact_deliverables": output.exists() and {p.name for p in output.iterdir()} == {"visual_brief.xlsx", "README.txt"},
              "readme_present": (output / "README.txt").is_file(),
              "workbook_present": workbook.is_file()}
    if openpyxl is None or not workbook.is_file():
        return {"pass": False, "score": 0, "checks": checks, "errors": ["workbook or openpyxl unavailable"]}
    try:
        with zipfile.ZipFile(workbook) as archive:
            checks["zip_integrity"] = archive.testzip() is None
        wb = openpyxl.load_workbook(workbook, data_only=False)
    except Exception as exc:
        return {"pass": False, "score": 0, "checks": checks, "errors": [f"workbook open failed: {type(exc).__name__}"]}
    chart_count = sum(len(ws._charts) for ws in wb.worksheets)
    # Count embedded images robustly: worksheet image objects via openpyxl AND any media
    # file present in the xlsx zip that is referenced by a drawing/sheet relationship.
    image_count = sum(len(ws._images) for ws in wb.worksheets)
    try:
        with zipfile.ZipFile(workbook) as archive:
            media_names = [n for n in archive.namelist() if n.startswith("xl/media/")]
            drawings = [n for n in archive.namelist() if n.startswith("xl/drawings/") and n.endswith(".xml")]
            # A drawing XML that references a blip means media is embedded in a worksheet.
            referenced = 0
            for d in drawings:
                xml = archive.read(d).decode("utf-8", errors="replace")
                if "<a:blip" in xml:
                    referenced += 1
            media_referenced = bool(drawings) and referenced > 0 and len(media_names) >= 2
            embedded_from_zip = max(image_count, len(media_names) if media_referenced else 0)
    except Exception:
        embedded_from_zip = image_count
    image_count = max(image_count, embedded_from_zip)
    checks["two_charts"] = chart_count >= 2
    checks["two_embedded_images"] = image_count >= 2
    checks["sources_sheet"] = "Sources" in wb.sheetnames
    candidates = {x.strip() for x in (input_dir / "image_candidates.txt").read_text().splitlines() if x.strip()}
    urls = []
    if "Sources" in wb.sheetnames:
        urls = [str(cell.value).strip() for row in wb["Sources"].iter_rows() for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("https://")]
    checks["candidate_urls"] = len(set(urls) & candidates) >= 2
    with (input_dir / "destination_performance.csv").open(newline="", encoding="utf-8") as handle:
        source = list(csv.reader(handle))
    raw_candidates = [ws for ws in wb.worksheets if ws.title.lower() in {"raw data", "data", "source data"}]
    checks["source_data_present"] = bool(raw_candidates) and raw_candidates[0].max_row >= len(source)
    checks["summary_present"] = any(ws.max_row >= 8 and ws.max_column >= 3 for ws in wb.worksheets if ws not in raw_candidates)
    weights = {"exact_deliverables": 10, "workbook_present": 10, "zip_integrity": 10,
               "two_charts": 20, "two_embedded_images": 20, "sources_sheet": 10,
               "candidate_urls": 10, "source_data_present": 5, "summary_present": 5, "readme_present": 5}
    score = sum(v for k, v in weights.items() if checks.get(k))
    passed = score >= 90 and all(checks.get(k) for k in ("exact_deliverables", "two_charts", "two_embedded_images", "candidate_urls", "source_data_present"))
    return {"pass": passed, "score": score, "checks": checks, "errors": [],
            "chart_count": chart_count, "image_count": image_count, "source_urls": sorted(set(urls))}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("output", type=Path)
    parser.add_argument("--input", type=Path, required=True)
    args = parser.parse_args()
    result = verify(args.output, args.input)
    print(json.dumps(result, indent=2, sort_keys=True))
    raise SystemExit(0 if result["pass"] else 1)


if __name__ == "__main__":
    main()

