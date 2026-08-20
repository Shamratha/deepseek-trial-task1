#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

from infra_lib import ROOT, tree_manifest, write_json


def workbook_summary(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        preview = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
            preview.append([v if isinstance(v, (str, int, float, bool)) or v is None else str(v) for v in row[:12]])
        sheets.append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column,
                       "charts": len(ws._charts), "images": len(ws._images), "preview": preview})
    return {"sheets": sheets}


def make_packet(run_dir: Path, label: str, rubric: Path) -> dict:
    output = run_dir / "output"
    artifacts = []
    for path in sorted(output.rglob("*")):
        if not path.is_file():
            continue
        item = {"path": path.relative_to(output).as_posix(), "bytes": path.stat().st_size}
        if path.suffix.lower() in {".txt", ".md", ".json"}:
            item["content"] = path.read_text(encoding="utf-8", errors="replace")[:50000]
        elif path.suffix.lower() == ".xlsx":
            item["workbook"] = workbook_summary(path)
        artifacts.append(item)
    return {"anonymous_label": label, "rubric": rubric.read_text(encoding="utf-8"),
            "artifacts": artifacts, "required_response": {"score": "0-100", "rationale": "string",
            "strengths": ["string"], "weaknesses": ["string"]}}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("run_dir", type=Path)
    parser.add_argument("--label", required=True)
    parser.add_argument("--rubric", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    write_json(args.output, make_packet(args.run_dir, args.label, args.rubric))


if __name__ == "__main__":
    main()

