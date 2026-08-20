from __future__ import annotations

import csv
import json
import os
import shutil
import sys
import tempfile
import time
from pathlib import Path

import openpyxl

import infra_lib
from graders.check1_verify import verify as verify_check1
from graders.check2_verify import verify as verify_check2
from graders.make_judge_packet import make_packet


ROOT = infra_lib.ROOT


def test_generators_are_deterministic():
    script = ROOT / "benchmarks/check1_text_excel/generate_input.py"
    os.system(f'"{sys.executable}" "{script}"')
    path = ROOT / "benchmarks/check1_text_excel/input/destination_performance.csv"
    first = infra_lib.sha256_file(path)
    os.system(f'"{sys.executable}" "{script}"')
    assert infra_lib.sha256_file(path) == first


def test_shared_datasets_are_byte_identical():
    one = ROOT / "benchmarks/check1_text_excel/input/destination_performance.csv"
    two = ROOT / "benchmarks/check2_visual_artifact/input/destination_performance.csv"
    assert one.read_bytes() == two.read_bytes()


def test_dataset_shape_and_seed_manifest():
    path = ROOT / "benchmarks/check1_text_excel/input/destination_performance.csv"
    with path.open(newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 144
    assert len({r["destination"] for r in rows}) == 12
    assert json.loads((path.parent.parent / "manifest.json").read_text())["generator_seed"] == 20260731


def test_workspace_copy_is_exact_and_isolated(tmp_path, monkeypatch):
    monkeypatch.setattr(infra_lib, "ROOT", tmp_path)
    monkeypatch.setattr(infra_lib, "CHECK_DIRS", {"check1": ROOT / "benchmarks/check1_text_excel"})
    workspace = infra_lib.prepare_workspace("check1", "test-run", infra_lib.ARMS[0])
    assert (workspace / "task.md").read_bytes() == (ROOT / "benchmarks/check1_text_excel/task.md").read_bytes()
    assert infra_lib.tree_manifest(workspace / "input") == infra_lib.tree_manifest(ROOT / "benchmarks/check1_text_excel/input")
    assert not (workspace / "destination_performance.csv").exists()


def test_redaction_hides_secret_named_values():
    assert infra_lib.redact({"api_key": "value", "safe": "ok"}) == {"api_key": "<REDACTED>", "safe": "ok"}


def test_codex_config_exact_slug():
    text = (ROOT / "config/codex-deepseek-config.toml").read_text()
    assert f'model = "{infra_lib.DEEPSEEK}"' in text
    assert 'model_provider = "openrouter"' in text


def test_claude_all_model_slots_pinned():
    env = infra_lib.sanitized_claude_env()
    for name in ("ANTHROPIC_DEFAULT_HAIKU_MODEL", "ANTHROPIC_DEFAULT_SONNET_MODEL",
                 "ANTHROPIC_DEFAULT_OPUS_MODEL", "CLAUDE_CODE_SUBAGENT_MODEL"):
        assert env[name] == infra_lib.DEEPSEEK


def test_runner_captures_process_fields(tmp_path):
    out, err = tmp_path / "out", tmp_path / "err"
    timing = infra_lib.run_process([sys.executable, "-c", "print('ok')"], tmp_path, dict(os.environ), 10, out, err)
    assert timing["exit_code"] == 0 and not timing["timed_out"]
    assert timing["wall_seconds"] >= 0 and out.read_text().strip() == "ok"


def test_timeout_kills_process_group(tmp_path):
    out, err = tmp_path / "out", tmp_path / "err"
    timing = infra_lib.run_process([sys.executable, "-c", "import time; time.sleep(30)"], tmp_path,
                                   dict(os.environ), 1, out, err)
    assert timing["timed_out"] and timing["exit_code"] != 0


def test_absent_telemetry_is_null(tmp_path):
    path = tmp_path / "events.jsonl"
    path.write_text('{"type":"done"}\n')
    usage = infra_lib.parse_usage_jsonl(path)
    assert usage["input_tokens"] is None and usage["reported_cost_usd"] is None


def test_reasoning_not_double_counted(tmp_path):
    path = tmp_path / "events.jsonl"
    path.write_text(json.dumps({"usage": {"input_tokens": 5, "output_tokens": 10,
                                          "reasoning_output_tokens": 4, "total_tokens": 15}}) + "\n")
    usage = infra_lib.parse_usage_jsonl(path)
    assert usage["total_tokens"] == 15 and usage["output_tokens"] == 10 and usage["reasoning_output_tokens"] == 4


def make_good_check1(output: Path):
    output.mkdir()
    (output / "report.txt").write_text("Management findings. " * 30)
    source = list(csv.reader((ROOT / "benchmarks/check1_text_excel/input/destination_performance.csv").open()))
    wb = openpyxl.Workbook()
    raw = wb.active
    raw.title = "Raw Data"
    for row in source:
        raw.append(row)
    dest = wb.create_sheet("Destination Summary")
    dest.append(["Destination", "Bookings", "Revenue"])
    for i, name in enumerate(sorted({r[1] for r in source[1:]}), 2):
        dest.append([name, f'=SUMIF(\'Raw Data\'!B:B,A{i},\'Raw Data\'!D:D)', f'=SUMIF(\'Raw Data\'!B:B,A{i},\'Raw Data\'!E:E)'])
    monthly = wb.create_sheet("Monthly Summary")
    monthly.append(["Month", "Bookings", "Revenue"])
    for i, month in enumerate(sorted({r[0] for r in source[1:]}), 2):
        monthly.append([month, f'=SUMIF(\'Raw Data\'!A:A,A{i},\'Raw Data\'!D:D)', f'=SUMIF(\'Raw Data\'!A:A,A{i},\'Raw Data\'!E:E)'])
    metrics = wb.create_sheet("Key Metrics")
    metrics.append(["Metric", "Value"])
    metrics.append(["Total bookings", "=SUM('Raw Data'!D2:D97)"])
    wb.save(output / "analysis.xlsx")


def test_check1_verifier_accepts_structurally_valid_workbook(tmp_path):
    output = tmp_path / "output"
    make_good_check1(output)
    result = verify_check1(output, ROOT / "benchmarks/check1_text_excel/input")
    assert result["pass"] and result["score"] >= 85


def test_check1_verifier_catches_malformed_workbook(tmp_path):
    output = tmp_path / "output"
    output.mkdir()
    (output / "report.txt").write_text("short")
    (output / "analysis.xlsx").write_text("not a zip")
    assert not verify_check1(output, ROOT / "benchmarks/check1_text_excel/input")["pass"]


def test_check2_verifier_detects_missing_charts_and_images(tmp_path):
    output = tmp_path / "output"
    output.mkdir()
    wb = openpyxl.Workbook()
    wb.save(output / "visual_brief.xlsx")
    result = verify_check2(output, ROOT / "benchmarks/check2_visual_artifact/input")
    assert not result["checks"]["two_charts"] and not result["checks"]["two_embedded_images"]


def test_judge_packet_is_anonymous(tmp_path):
    run_dir = tmp_path / "run"
    (run_dir / "output").mkdir(parents=True)
    (run_dir / "output/report.txt").write_text("artifact")
    rubric = tmp_path / "rubric.md"
    rubric.write_text("Score quality")
    packet = make_packet(run_dir, "B", rubric)
    assert packet["anonymous_label"] == "B" and "control_native_codex" not in json.dumps(packet)


def test_result_schema_stable():
    assert len(infra_lib.METRIC_FIELDS) == len(set(infra_lib.METRIC_FIELDS))
    assert {"check_id", "arm", "total_tokens", "judge_score", "notes"}.issubset(infra_lib.METRIC_FIELDS)


def test_secret_scanner_catches_fake_fixture(tmp_path):
    token = "sk-" + "or-v1-" + "x" * 20
    (tmp_path / "fixture.txt").write_text(token)
    assert infra_lib.scan_secrets([tmp_path])


def test_check3_remains_blocked():
    assert not infra_lib.check3_ready()


def test_control_rejects_custom_provider(tmp_path):
    (tmp_path / "auth.json").write_text("{}")
    (tmp_path / "config.toml").write_text('model_provider="openrouter"')
    assert infra_lib.control_config_errors({"CONTROL_CODEX_HOME": str(tmp_path)})


def test_command_manifest_contains_no_secret_values(tmp_path):
    command = ["tool", "prompt"]
    child = {"OPENROUTER_API_KEY": "sensitive", "PATH": "/bin"}
    value = infra_lib.command_manifest(command, child)
    assert "sensitive" not in json.dumps(value) and "OPENROUTER_API_KEY" in value["environment_variable_names"]
