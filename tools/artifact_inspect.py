#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser(description="Read-only workbook structure inspector")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    wb = openpyxl.load_workbook(args.workbook, read_only=False, data_only=False)
    result = {"path": str(args.workbook), "sheets": []}
    for ws in wb.worksheets:
        formulas = sum(1 for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("="))
        result["sheets"].append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column,
                                 "charts": len(ws._charts), "images": len(ws._images), "formulas": formulas})
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
